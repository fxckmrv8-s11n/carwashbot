"""
Генератор отчёта в формате .xlsx — для бухгалтерии.
Использует те же данные, что и pdf_generator.generate_pdf (day_data, summary).
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="1C1C1E")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TOTAL_FONT = Font(bold=True)
THIN = Side(style="thin", color="DDDDDD")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _autosize(ws):
    for col_cells in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = max(10, length + 2)


def generate_xlsx(day_data: dict, summary: dict, output_path: str) -> None:
    wb = Workbook()

    # ── Лист "Машины" ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Машины"
    headers = ["№", "Время", "Машина", "Мойщик", "Услуги", "Оплата", "Сумма, ₽"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER

    for c in day_data.get("cars", []):
        payment = c.get("payment", "")
        if c.get("payment_split"):
            payment = " + ".join(f"{k} {v}₽" for k, v in c["payment_split"].items())
        ws.append([
            c.get("num", ""), c.get("time", ""), c.get("car", "") or "—",
            c.get("employee", ""), c.get("service", ""), payment, c.get("price", 0),
        ])
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = BORDER
    _autosize(ws)

    # ── Лист "Итоги" ───────────────────────────────────────────────────
    ws2 = wb.create_sheet("Итоги")
    ws2.append(["Показатель", "Значение"])
    for cell in ws2[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER
    rows = [
        ("Дата", day_data.get("date", "")),
        ("Всего машин", len(day_data.get("cars", []))),
        ("Итого, ₽", summary.get("grand_total", summary.get("total", 0))),
        ("Наличные, ₽", summary.get("cash", 0)),
        ("Безнал, ₽", summary.get("beznal", 0)),
        ("Карта, ₽", summary.get("visa", 0)),
    ]
    for r in rows:
        ws2.append(list(r))
    for row in ws2.iter_rows(min_row=2):
        for cell in row:
            cell.border = BORDER
    ws2["B3"].font = TOTAL_FONT
    _autosize(ws2)

    # ── Лист "Зарплаты" ────────────────────────────────────────────────
    ws3 = wb.create_sheet("Зарплаты")
    ws3.append(["Мойщик", "Машин", "К выплате, ₽"])
    for cell in ws3[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER
    counts = {}
    for c in day_data.get("cars", []):
        counts[c.get("employee", "—")] = counts.get(c.get("employee", "—"), 0) + 1
    for emp, salary in (summary.get("washer_salaries") or {}).items():
        ws3.append([emp, counts.get(emp, 0), salary])
    for row in ws3.iter_rows(min_row=2):
        for cell in row:
            cell.border = BORDER
    _autosize(ws3)

    wb.save(output_path)
